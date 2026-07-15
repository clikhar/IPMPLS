"""In-memory CSV, Excel, and PDF reporting without writing sensitive files to disk."""
import csv
from io import BytesIO, StringIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session
from crtnm.infrastructure.models import AuditLogModel, DeviceModel, StationModel


class ReportService:
    """Exports inventory and audit evidence using only non-secret fields."""

    INVENTORY_HEADERS = ["Station", "Division", "Device", "Type", "Vendor", "Model", "Management IP", "Protocol"]

    @classmethod
    def inventory_rows(cls, session: Session) -> list[list[str]]:
        """Return reporting rows without usernames, passwords, or cipher text."""
        statement = select(StationModel.name, StationModel.division, DeviceModel.name, DeviceModel.device_type, DeviceModel.vendor, DeviceModel.model, DeviceModel.management_ip, DeviceModel.protocol).join(DeviceModel, DeviceModel.station_id == StationModel.id).order_by(StationModel.name, DeviceModel.name)
        return [[str(value or "") for value in row] for row in session.execute(statement).all()]

    @staticmethod
    def to_csv(headers: list[str], rows: list[list[str]]) -> bytes:
        """Create UTF-8 CSV bytes suitable for a download response."""
        stream = StringIO(); writer = csv.writer(stream); writer.writerow(headers); writer.writerows(rows)
        return stream.getvalue().encode("utf-8-sig")

    @staticmethod
    def to_xlsx(headers: list[str], rows: list[list[str]]) -> bytes:
        """Build a formatted standalone workbook in memory."""
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Device Inventory"; sheet.append(headers)
        for row in rows: sheet.append(row)
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="17365D")
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(40, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        stream = BytesIO(); workbook.save(stream); return stream.getvalue()

    @classmethod
    def inventory_pdf(cls, session: Session) -> bytes:
        """Build a compact printable inventory report in memory."""
        stream = BytesIO(); document = SimpleDocTemplate(stream, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=28, bottomMargin=28)
        styles = getSampleStyleSheet(); rows = cls.inventory_rows(session)
        table = Table([cls.INVENTORY_HEADERS, *rows], repeatRows=1, colWidths=[60, 60, 72, 55, 52, 60, 70, 45])
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C4CE")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F9")])]))
        document.build([Paragraph("CRTNM Device Inventory", styles["Title"]), Spacer(1, 12), table])
        return stream.getvalue()

    @staticmethod
    def audit_rows(session: Session, limit: int = 500) -> list[list[str]]:
        """Return the most recent audit records for viewing/export."""
        events = session.scalars(select(AuditLogModel).order_by(AuditLogModel.id.desc()).limit(limit))
        return [[str(event.created_at), event.actor, event.action, event.target, event.detail or ""] for event in events]
